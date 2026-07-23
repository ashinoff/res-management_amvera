#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Анализатор профиля мощности (выгрузка «ПРОФИЛЬ МОЩНОСТИ ДЛЯ 1С» из Пирамиды).

Контракт (как у nartis/energomera): argv[1] = путь к .xlsx, результат — JSON в stdout.
Листы «30 мин» и «60 мин», одинаковая шапка:
  строка 5: «№ прибора учёта» — номера ПУ по колонкам начиная с C;
  строка 6: «Ктт/Ктн» вида «200/1» → kt = Ктт × Ктн;
  строка 8: заголовки; данные со строки 9 до строки «Итого…»;
  колонка A: дата dd.mm.yyyy; колонка B: время HH:MM («24:00» = конец суток →
  трактуем как 00:00 следующего дня при построении метки времени).

Методика (зафиксирована заказчиком, не менять):
 1. Часовой ряд: если на «60 мин» есть значения — берём как есть. Иначе по «30 мин»:
    час H:00 = (получасовка H:30 + получасовка (H+1):00) / 2; отсутствующая = 0.
 2. Пик: max часового ряда; peakKw = max × kt; фиксируем dateTime пика.
 3. Энергия: сумма ряда × kt.
 4. Период: min/max дат файла.
"""
import json
import sys
from datetime import datetime, timedelta

import openpyxl


PU_ROW = 5        # строка с номерами ПУ
KT_ROW = 6        # строка Ктт/Ктн
DATA_START = 9    # первая строка данных
FIRST_PU_COL = 3  # колонка C


def _num(v):
    """Число из ячейки или None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '.').replace('\xa0', '').replace(' ', '')
    if s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _pu_str(v):
    """Номер ПУ в строку без хвоста '.0' (openpyxl отдаёт числа как float)."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    # «1294249.0» → «1294249»
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s


def _parse_kt(v):
    """«200/1» → 200.0 (Ктт×Ктн). При невозможности — 1.0."""
    if v is None:
        return 1.0
    s = str(v).strip()
    parts = [p for p in s.replace('\\', '/').split('/') if p.strip() != '']
    prod = 1.0
    found = False
    for p in parts:
        n = _num(p)
        if n is not None:
            prod *= n
            found = True
    return prod if found and prod != 0 else 1.0


def _parse_date(v):
    """Дата из ячейки A (datetime или строка dd.mm.yyyy)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if s == '':
        return None
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_time_parts(v):
    """Время из ячейки B → (hour, minute, is_2400). «24:00» → (0,0,True)."""
    if v is None:
        return None
    # datetime.time
    if hasattr(v, 'hour') and hasattr(v, 'minute') and not isinstance(v, str):
        return (v.hour, v.minute, False)
    s = str(v).strip()
    if s == '':
        return None
    if s.startswith('24:'):
        return (0, 0, True)
    try:
        parts = s.split(':')
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        if h == 24:
            return (0, 0, True)
        return (h, m, False)
    except (ValueError, IndexError):
        return None


def _cell_dt(date_val, time_val):
    """Строит datetime метки. «24:00» → 00:00 следующего дня."""
    d = _parse_date(date_val)
    t = _parse_time_parts(time_val)
    if d is None or t is None:
        return None
    hour, minute, is_2400 = t
    dt = datetime(d.year, d.month, d.day, hour, minute)
    if is_2400:
        dt = dt + timedelta(days=1)
    return dt


def _find_sheet(wb, needle):
    for name in wb.sheetnames:
        n = name.lower().replace(' ', '')
        if needle in n and 'мин' in name.lower():
            return wb[name]
    return None


def _read_sheet(ws):
    """Возвращает {col_index: {'pu': str, 'kt': float}} и {col_index: {dt: kw_raw}}.
    dt — метка времени (24:00 → следующие сутки), kw_raw — значение без kt.

    ОДИН проход через iter_rows(values_only=True): случайный ws.cell(row,col) в
    openpyxl (особенно read-only) перепарсивает лист на каждый вызов → квадратично
    на реальном файле (1449×8). Здесь строки читаются последовательно, метаданные
    (ПУ/Ктт·Ктн) берём из строк 5/6, данные — с 9-й до «Итого»."""
    if ws is None:
        return {}, {}, []

    meta = {}                 # col -> {'pu': str, 'kt': float}
    series = {}               # col -> {dt: kw_raw}
    dates = []
    pu_row_vals = None
    kt_row_vals = None

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == PU_ROW:
            pu_row_vals = row
            continue
        if idx == KT_ROW:
            kt_row_vals = row
            continue
        if idx < DATA_START:
            continue

        # Первая строка данных — строим карту колонок из уже прочитанных шапок.
        if not meta and pu_row_vals is not None:
            ncols = max(len(pu_row_vals), len(kt_row_vals) if kt_row_vals else 0)
            for col in range(FIRST_PU_COL, ncols + 1):
                pu_cell = pu_row_vals[col - 1] if col - 1 < len(pu_row_vals) else None
                pu_s = _pu_str(pu_cell)
                if pu_s:
                    kt_cell = kt_row_vals[col - 1] if (kt_row_vals and col - 1 < len(kt_row_vals)) else None
                    meta[col] = {'pu': pu_s, 'kt': _parse_kt(kt_cell)}
                    series[col] = {}

        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        a_s = str(a).strip().lower() if a is not None else ''
        if a_s.startswith('итого'):
            break
        dt = _cell_dt(a, b)
        if dt is not None:
            d = _parse_date(a)
            if d is not None:
                dates.append(d)
            for c in meta:
                val = _num(row[c - 1]) if c - 1 < len(row) else None
                if val is not None:
                    series[c][dt] = val

    return meta, series, dates


def _hourly_from_60(dt_map):
    """Часовой ряд из 60-мин: берём значения как есть (ключ — метка часа)."""
    return dict(dt_map)


def _hourly_from_30(dt_map):
    """Часовой ряд из 30-мин: час H:00 = (H:30 + (H+1):00)/2, пропуск = 0."""
    hourly = {}
    hour_starts = sorted([dt for dt in dt_map if dt.minute == 0])
    for h in hour_starts:
        half = dt_map.get(h + timedelta(minutes=30), 0.0)
        nexth = dt_map.get(h + timedelta(minutes=60), 0.0)
        hourly[h] = (half + nexth) / 2.0
    return hourly


def analyze(filepath):
    # НЕ read_only: файлы небольшие (обычная загрузка ~0.5 c), а read-only делает
    # случайный доступ к ячейкам квадратичным. Данные читаем одним проходом.
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws60 = _find_sheet(wb, '60')
    ws30 = _find_sheet(wb, '30')

    meta60, series60, dates60 = _read_sheet(ws60)
    meta30, series30, dates30 = _read_sheet(ws30)

    # Полный список ПУ (объединение по номерам) с их kt (приоритет — где есть данные)
    pu_info = {}   # pu -> {'kt': float}
    for m in (meta60, meta30):
        for c, info in m.items():
            pu_info.setdefault(info['pu'], {})['kt'] = info['kt']

    all_dates = dates60 + dates30
    period = ''
    if all_dates:
        period = f"{min(all_dates).strftime('%d.%m.%Y')}–{max(all_dates).strftime('%d.%m.%Y')}"

    # col по ПУ на каждом листе
    col60 = {info['pu']: c for c, info in meta60.items()}
    col30 = {info['pu']: c for c, info in meta30.items()}

    results = []
    warnings = []

    for pu, info in pu_info.items():
        kt = info.get('kt', 1.0)
        hourly = {}
        source = None

        if pu in col60 and series60.get(col60[pu]):
            hourly = _hourly_from_60(series60[col60[pu]])
            source = '60'
        elif pu in col30 and series30.get(col30[pu]):
            hourly = _hourly_from_30(series30[col30[pu]])
            source = '30'

        if not hourly:
            warnings.append(pu)
            continue

        # Пик
        peak_dt = max(hourly, key=lambda d: hourly[d])
        peak_raw = hourly[peak_dt]
        peak_kw = peak_raw * kt
        energy_kwh = sum(hourly.values()) * kt

        results.append({
            'puNumber': pu,
            'kt': round(kt, 4),
            'peakRaw': round(peak_raw, 4),
            'peakKw': round(peak_kw, 4),
            'peakAt': peak_dt.strftime('%d.%m.%Y %H:%M'),
            'energyKwh': round(energy_kwh, 4),
            'source': source,
            'period': period,
        })

    return {
        'success': True,
        'results': results,
        'warnings': warnings,
        'period': period,
    }


if __name__ == '__main__':
    # UTF-8 в stdout независимо от locale контейнера (иначе en-dash в period при
    # POSIX/C-локали упал бы с UnicodeEncodeError и Node получил бы пустой вывод).
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass
    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': 'No file path provided'}))
        sys.exit(1)
    try:
        print(json.dumps(analyze(sys.argv[1]), ensure_ascii=False))
    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False))
